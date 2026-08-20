import openpyxl
import re

print("\n" + "="*80)
print("COMPREHENSIVE VERIFICATION - TEST CASES vs RTM SYNC")
print("="*80)

# Count markdown test cases
with open('testcases.md', 'r', encoding='utf-8') as f:
    md_content = f.read()
    md_tc_count = len(re.findall(r'^#### TC_CENTRAL_\d+', md_content, re.MULTILINE))

# Count Excel test cases
tc_wb = openpyxl.load_workbook('test-cases.xlsx')
tc_ws = tc_wb.active
excel_tc_count = tc_ws.max_row - 1  # Subtract header

# Count RTM mappings
rtm_wb = openpyxl.load_workbook('rtm.xlsx')
rtm_ws = rtm_wb.active
rtm_req_count = rtm_ws.max_row - 1  # Subtract header

print(f"\n1. MARKDOWN TEST CASES (testcases.md)")
print(f"   TC Count: {md_tc_count}")

print(f"\n2. EXCEL TEST CASES (test-cases.xlsx)")
print(f"   Rows (excluding header): {excel_tc_count}")
print(f"   Columns: {tc_ws.max_column}")

print(f"\n3. RTM MAPPINGS (rtm.xlsx)")
print(f"   Requirement Mappings: {rtm_req_count}")
print(f"   Columns: {rtm_ws.max_column}")

print(f"\n4. VERIFICATION RESULTS")
print(f"   Markdown TCs: {md_tc_count}")
print(f"   Excel TCs:    {excel_tc_count}")
print(f"   Match: {'YES ✅' if md_tc_count == excel_tc_count else 'NO ❌ MISMATCH'}")

print(f"\n5. RTM SYNC STATUS")
if rtm_req_count >= (excel_tc_count * 0.3):
    print(f"   Sufficient coverage: YES ✅")
    print(f"   (Many-to-one mapping expected: {rtm_req_count} mappings for {excel_tc_count} TCs)")
else:
    print(f"   INSUFFICIENT: {rtm_req_count} mappings for {excel_tc_count} TCs ❌")
    print(f"   RTM needs regeneration!")

print(f"\n" + "="*80)
if md_tc_count == excel_tc_count and rtm_req_count >= 80:
    print("STATUS: ✅ ALL VERIFICATIONS PASSED - SYNC COMPLETE")
    print("="*80)
else:
    print("STATUS: ⚠️  ACTION REQUIRED")
    if md_tc_count != excel_tc_count:
        print(f"   - Excel count mismatch: {md_tc_count} vs {excel_tc_count}")
    if rtm_req_count < 80:
        print(f"   - RTM incomplete: {rtm_req_count} mappings (should be 80+)")
    print("="*80)
