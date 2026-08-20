#!/usr/bin/env python3
"""
Convert testcases_v4.0_quality_focused.md to test-cases.xlsx
Implements Constitutional Rule VI: File Synchronization
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_markdown_testcases(md_file):
    """Parse test cases from markdown file"""
    with open(md_file, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
    
    test_cases = []
    
    # Split by test case headers (### TC_CENTRAL_XXX:)
    tc_blocks = re.split(r'### (TC_CENTRAL_\d+):\s*(.+?)\n', content)
    
    # Process test case blocks
    i = 1
    while i < len(tc_blocks):
        if i + 2 >= len(tc_blocks):
            break
            
        tc_id = tc_blocks[i].strip()
        tc_summary = tc_blocks[i + 1].strip()
        tc_content = tc_blocks[i + 2] if i + 2 < len(tc_blocks) else ""
        
        # Extract fields from table
        fields = {
            'Test Case ID': tc_id,
            'Summary': tc_summary,
            'Requirement': '',
            'Prerequisites': '',
            'Steps': '',
            'Expected Output': '',
            'Priority': '',
            'Severity': '',
            'Category': ''
        }
        
        # Parse the markdown table for this test case
        # Look for | **Field Name** | Value |
        table_pattern = r'\|\s*\*\*(.+?)\*\*\s*\|\s*(.+?)\s*\|'
        table_matches = re.findall(table_pattern, tc_content)
        
        for key, value in table_matches:
            key = key.strip()
            value = value.strip()
            # Remove markdown formatting
            value = value.replace('**', '').replace('`', '')
            
            if key == 'Test Case ID':
                fields['Test Case ID'] = value
            elif key == 'Requirement':
                fields['Requirement'] = value
            elif key == 'Summary':
                fields['Summary'] = value
            elif key == 'Prerequisites':
                fields['Prerequisites'] = value
            elif key == 'Steps':
                fields['Steps'] = value
            elif key == 'Expected Output':
                fields['Expected Output'] = value
            elif key == 'Priority':
                fields['Priority'] = value
            elif key == 'Severity':
                fields['Severity'] = value
            elif key == 'Category':
                fields['Category'] = value
        
        if fields['Test Case ID']:
            test_cases.append(fields)
        
        i += 3
    
    return test_cases

def create_excel_from_testcases(test_cases, output_file):
    """Create Excel workbook with test cases"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Define columns (Constitutional Rule VII: 12-Column Format)
    columns = [
        'Test Case ID',
        'Summary',
        'Requirement',
        'Prerequisites',
        'Steps',
        'Expected Output',
        'Actual Result',
        'Priority',
        'Severity',
        'Category',
        'Status',
        'Comments'
    ]
    
    # Add header row
    for col_num, col_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = col_title
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Set column widths
    column_widths = {
        'A': 15,  # Test Case ID
        'B': 40,  # Summary
        'C': 15,  # Requirement
        'D': 30,  # Prerequisites
        'E': 30,  # Steps
        'F': 30,  # Expected Output
        'G': 20,  # Actual Result
        'H': 10,  # Priority
        'I': 12,  # Severity
        'J': 15,  # Category
        'K': 12,  # Status
        'L': 25   # Comments
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add test cases
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_num, tc in enumerate(test_cases, 2):
        ws.cell(row=row_num, column=1).value = tc.get('Test Case ID', '')
        ws.cell(row=row_num, column=2).value = tc.get('Summary', '')
        ws.cell(row=row_num, column=3).value = tc.get('Requirement', '')
        ws.cell(row=row_num, column=4).value = tc.get('Prerequisites', '')
        ws.cell(row=row_num, column=5).value = tc.get('Steps', '')
        ws.cell(row=row_num, column=6).value = tc.get('Expected Output', '')
        ws.cell(row=row_num, column=7).value = ''  # Actual Result (empty - for QA to fill)
        ws.cell(row=row_num, column=8).value = tc.get('Priority', '')
        ws.cell(row=row_num, column=9).value = tc.get('Severity', '')
        ws.cell(row=row_num, column=10).value = tc.get('Category', '')
        ws.cell(row=row_num, column=11).value = 'Not Started'  # Status
        ws.cell(row=row_num, column=12).value = ''  # Comments
        
        # Apply borders and alignment
        for col in range(1, 13):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save(output_file)
    print(f"✅ Excel file created: {output_file}")
    print(f"✅ Total test cases: {len(test_cases)}")
    print(f"✅ Format: 12-Column Constitutional Rule VII")

if __name__ == "__main__":
    md_file = "testcases_v4.0_quality_focused.md"
    output_file = "test-cases.xlsx"
    
    print(f"📖 Reading test cases from: {md_file}")
    test_cases = parse_markdown_testcases(md_file)
    print(f"📊 Parsed {len(test_cases)} test cases")
    
    print(f"\n📝 Creating Excel file: {output_file}")
    create_excel_from_testcases(test_cases, output_file)
    
    print(f"\n✅ SYNC COMPLETE: Constitutional Rule VI (File Synchronization) ✅")
    print(f"   Markdown: {md_file} ({len(test_cases)} TCs)")
    print(f"   Excel: {output_file} ({len(test_cases)} TCs)")
    print(f"   Status: ✅ SYNCHRONIZED")
